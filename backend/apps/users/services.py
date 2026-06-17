"""
User app service layer — Group management + User CRUD + bulk import.

All business logic lives here; views only handle HTTP serialisation.
"""
from __future__ import annotations

import io
import secrets
import string
from typing import TYPE_CHECKING, TypedDict

from django.conf import settings
from django.core.exceptions import ValidationError
from django.core.mail import send_mail
from django.core.validators import validate_email as _validate_email_format

from apps.reports.models import AuditLog

from .models import Area, Group, User, UserProfile

if TYPE_CHECKING:
    from django.core.files.uploadedfile import UploadedFile


# ---------------------------------------------------------------------------
# Bulk-import types
# ---------------------------------------------------------------------------

MAX_IMPORT_FILE_BYTES = 5 * 1024 * 1024  # 5 MB

XLSX_MAGIC = b"PK\x03\x04"  # All .xlsx / OpenXML files are ZIP archives

# Maps Spanish labels (and English codes) to internal role codes
_ROLE_MAP: dict[str, str] = {
    "administrador": User.Role.ADMIN,
    "admin": User.Role.ADMIN,
    "capacitador": User.Role.TRAINER,
    "trainer": User.Role.TRAINER,
    "usuario": User.Role.USUARIO,
}


class ValidRow(TypedDict):
    row: int
    email: str
    first_name: str
    last_name: str
    role: str
    area_id: int | None
    cargo: str
    grupo_id: int | None


class ErrorRow(TypedDict):
    row: int
    email: str
    errors: list[str]


class BulkImportPreviewResult(TypedDict):
    valid_count: int
    error_count: int
    valid_rows: list[ValidRow]
    error_rows: list[ErrorRow]


class BulkImportCommitResult(TypedDict):
    created: int
    failed: int
    errors: list[ErrorRow]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _generate_temp_password() -> str:
    """
    Generate a 12-character password guaranteed to satisfy the policy:
    uppercase, lowercase, digit, special char.
    """
    special = "!@#$%^&*"
    pool = string.ascii_letters + string.digits + special
    pwd: list[str] = [
        secrets.choice(string.ascii_uppercase),
        secrets.choice(string.ascii_lowercase),
        secrets.choice(string.digits),
        secrets.choice(special),
    ] + [secrets.choice(pool) for _ in range(8)]
    secrets.SystemRandom().shuffle(pwd)
    return "".join(pwd)


def _blacklist_all_user_tokens(user: User) -> None:
    """Blacklist every outstanding JWT refresh token for this user."""
    try:
        from rest_framework_simplejwt.token_blacklist.models import (
            BlacklistedToken,
            OutstandingToken,
        )

        for token in OutstandingToken.objects.filter(user=user):
            BlacklistedToken.objects.get_or_create(token=token)
    except Exception:  # noqa: BLE001
        pass


def _username_from_email(email: str) -> str:
    """Derive a unique username from an email address."""
    base = email.split("@")[0][:20]
    candidate = base
    counter = 1
    while User.objects.filter(username=candidate).exists():
        candidate = f"{base}{counter}"
        counter += 1
    return candidate


def _resolve_area(area_name: str) -> "Area | None":
    """Look up an active Area by name (case-insensitive). Returns None if not found."""
    if not area_name:
        return None
    return Area.objects.filter(nombre__iexact=area_name, activo=True).first()


# ---------------------------------------------------------------------------
# User CRUD
# ---------------------------------------------------------------------------


def create_user(
    *,
    email: str,
    first_name: str,
    last_name: str,
    role: str = User.Role.USUARIO,
    area: str = "",
    area_id: int | None = None,
    cargo: str = "",
    grupo_id: int | None = None,
    admin_user: User,
    ip: str,
) -> User:
    """
    Create a new user with a temporary password and send a welcome email.

    * The user is created with must_change_password=True.
    * An email with the temporary password is sent immediately.
    * An AuditLog entry is created by the admin.
    """
    temp_password = _generate_temp_password()
    username = _username_from_email(email)

    user = User.objects.create_user(
        username=username,
        email=email,
        first_name=first_name,
        last_name=last_name,
        password=temp_password,
        role=role,
        must_change_password=True,
    )

    # Update the auto-created profile (signal created it with blank fields)
    profile = user.profile
    if area_id is not None:
        profile.area_id = area_id
    else:
        profile.area = _resolve_area(area)
    profile.cargo = cargo
    if grupo_id is not None:
        profile.grupo_id = grupo_id
    profile.save()

    # Send welcome email
    frontend_url = getattr(settings, "FRONTEND_URL", "http://localhost:3000")
    from_email = settings.DEFAULT_FROM_EMAIL or getattr(settings, "EMAIL_HOST_USER", "")
    send_mail(
        subject="Bienvenido al Sistema de Capacitaciones",
        message=(
            f"Hola {first_name},\n\n"
            f"Se ha creado tu cuenta en el sistema de capacitaciones.\n\n"
            f"Correo: {email}\n"
            f"Contraseña temporal: {temp_password}\n\n"
            f"Por seguridad, deberás cambiar tu contraseña al iniciar sesión por primera vez.\n\n"
            f"Ingresa aquí: {frontend_url}/login\n\n"
            f"Si no esperabas este mensaje, ignóralo."
        ),
        from_email=from_email,
        recipient_list=[email],
        fail_silently=True,
    )

    AuditLog.objects.create(
        user=admin_user,
        accion="USER_CREATED",
        ip=ip,
        detalles_json={"target_user_id": user.pk, "email": email, "role": role},
    )

    return user


def update_user(
    user: User,
    *,
    first_name: str | None = None,
    last_name: str | None = None,
    area: str | None = None,
    area_id: int | None = None,
    cargo: str | None = None,
    grupo_id: int | None = None,
) -> User:
    """Update allowed user and profile fields (partial update)."""
    user_changed = False
    if first_name is not None:
        user.first_name = first_name
        user_changed = True
    if last_name is not None:
        user.last_name = last_name
        user_changed = True
    if user_changed:
        user.save(update_fields=["first_name", "last_name"])

    profile = user.profile
    profile_changed = False
    if area_id is not None:
        profile.area_id = area_id
        profile_changed = True
    elif area is not None:
        profile.area = _resolve_area(area)
        profile_changed = True
    if cargo is not None:
        profile.cargo = cargo
        profile_changed = True
    if grupo_id is not None:
        profile.grupo_id = grupo_id
        profile_changed = True
    if profile_changed:
        profile.save()

    return user


def change_role(user: User, *, new_role: str, admin_user: User, ip: str) -> User:
    """Change a user's role and record the event in AuditLog."""
    old_role = user.role
    user.role = new_role
    user.save(update_fields=["role"])

    AuditLog.objects.create(
        user=admin_user,
        accion="ROLE_CHANGED",
        ip=ip,
        detalles_json={
            "target_user_id": user.pk,
            "old_role": old_role,
            "new_role": new_role,
        },
    )
    return user


def deactivate_user(user: User, *, admin_user: User, ip: str) -> User:
    """
    Deactivate a user account:
    1. Set is_active=False
    2. Immediately blacklist ALL outstanding refresh tokens (closes active sessions)
    3. Record in AuditLog
    """
    user.is_active = False
    user.save(update_fields=["is_active"])

    _blacklist_all_user_tokens(user)

    AuditLog.objects.create(
        user=admin_user,
        accion="USER_DEACTIVATED",
        ip=ip,
        detalles_json={"target_user_id": user.pk},
    )
    return user


# ---------------------------------------------------------------------------
# Group CRUD
# ---------------------------------------------------------------------------


def create_group(*, nombre: str, descripcion: str = "", activo: bool = True) -> Group:
    return Group.objects.create(nombre=nombre, descripcion=descripcion, activo=activo)


def update_group(group: Group, **kwargs: object) -> Group:
    for field, value in kwargs.items():
        setattr(group, field, value)
    group.save()
    return group


def delete_group(group: Group) -> None:
    """Delete a group.  Raises ValidationError if the group still has members."""
    if group.members.exists():
        raise ValidationError(
            {"non_field_errors": ["No se puede eliminar un grupo que tiene miembros activos."]}
        )
    group.delete()


# ---------------------------------------------------------------------------
# Member management
# ---------------------------------------------------------------------------


def add_members(group: Group, user_ids: list[int]) -> list[UserProfile]:
    """Assign users to *group*.  Returns the updated UserProfile instances."""
    profiles: list[UserProfile] = []
    for uid in user_ids:
        try:
            profile = UserProfile.objects.select_related("user").get(user_id=uid)
        except UserProfile.DoesNotExist:
            raise ValidationError(
                {"user_ids": [f"No existe un usuario con id={uid}."]}
            )
        profile.grupo = group
        profile.save()
        profiles.append(profile)
    return profiles


def remove_member(group: Group, user_id: int) -> None:
    """Remove *user_id* from *group*.  Raises ValidationError if not a member."""
    try:
        profile = UserProfile.objects.get(user_id=user_id, grupo=group)
    except UserProfile.DoesNotExist:
        raise ValidationError(
            {"user_id": [f"El usuario id={user_id} no pertenece a este grupo."]}
        )
    profile.grupo = None
    profile.save()


# ---------------------------------------------------------------------------
# Bulk import
# ---------------------------------------------------------------------------


def bulk_import_preview(
    file: "UploadedFile",
) -> BulkImportPreviewResult:
    """
    Parse an uploaded .xlsx file and return a preview of valid/invalid rows.

    Does NOT create any users. Validates:
    - Magic bytes (must be XLSX / ZIP format)
    - File size (≤ 5 MB)
    - Required columns: email, nombre, apellido, rol
    - Per-row: email format, email uniqueness, role value, non-empty name fields
    - Optional: area, cargo, grupo (looked up by name)
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError as exc:
        raise ValidationError({"file": ["El servidor no tiene soporte para archivos Excel."]}) from exc

    # ── File size guard ──────────────────────────────────────────────────────
    file.seek(0, 2)  # seek to end
    size = file.tell()
    file.seek(0)
    if size > MAX_IMPORT_FILE_BYTES:
        raise ValidationError(
            {"file": [f"El archivo supera el límite de {MAX_IMPORT_FILE_BYTES // 1024 // 1024} MB."]}
        )

    # ── Magic bytes ─────────────────────────────────────────────────────────
    magic = file.read(4)
    file.seek(0)
    if magic[:4] != XLSX_MAGIC:
        raise ValidationError(
            {"file": ["El archivo debe estar en formato Excel (.xlsx)."]}
        )

    # ── Parse workbook ──────────────────────────────────────────────────────
    try:
        wb = load_workbook(
            filename=io.BytesIO(file.read()),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValidationError({"file": ["No se pudo leer el archivo Excel. Verifica que no esté corrupto."]}) from exc

    ws = wb.active
    if ws is None:
        raise ValidationError({"file": ["El archivo Excel no contiene hojas de cálculo."]})

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValidationError({"file": ["El archivo está vacío."]})

    # ── Build column index from header row ──────────────────────────────────
    header = [str(c).strip().lower() if c else "" for c in all_rows[0]]
    REQUIRED_COLS = ["email", "nombre", "apellido", "rol"]
    col_idx: dict[str, int] = {}
    missing: list[str] = []
    for col in REQUIRED_COLS:
        try:
            col_idx[col] = header.index(col)
        except ValueError:
            missing.append(col)
    if missing:
        raise ValidationError(
            {"file": [f"Faltan columnas obligatorias en el encabezado: {', '.join(missing)}"]}
        )

    # Optional columns
    for col in ["area", "cargo", "grupo"]:
        try:
            col_idx[col] = header.index(col)
        except ValueError:
            col_idx[col] = -1  # not present

    # Pre-load group names for fast lookup
    group_by_name: dict[str, int] = {
        name.lower(): gid
        for gid, name in Group.objects.values_list("id", "nombre")
    }

    # Pre-load active area names for fast lookup
    area_by_name: dict[str, int] = {
        name.lower(): aid
        for aid, name in Area.objects.filter(activo=True).values_list("id", "nombre")
    }

    # Track emails seen so far in *this file* to catch intra-file duplicates
    seen_emails: set[str] = set()

    valid_rows: list[ValidRow] = []
    error_rows: list[ErrorRow] = []

    def _cell(row_values: tuple, col: str) -> str:
        idx = col_idx.get(col, -1)
        if idx < 0 or idx >= len(row_values):
            return ""
        return str(row_values[idx]).strip() if row_values[idx] is not None else ""

    for row_number, row_values in enumerate(all_rows[1:], start=2):
        # Skip completely empty rows
        if all(v is None or str(v).strip() == "" for v in row_values):
            continue

        email_raw = _cell(row_values, "email")
        errors: list[str] = []

        # ── Email ────────────────────────────────────────────────────────────
        email = email_raw.lower().strip()
        if not email:
            errors.append("El correo electrónico es obligatorio.")
        else:
            try:
                _validate_email_format(email)
            except Exception:
                errors.append("El correo electrónico no tiene un formato válido.")
            else:
                if email in seen_emails:
                    errors.append("Correo electrónico duplicado dentro del archivo.")
                elif User.objects.filter(email=email).exists():
                    errors.append("Ya existe un usuario con este correo.")
                else:
                    seen_emails.add(email)

        # ── First / last name ────────────────────────────────────────────────
        first_name = _cell(row_values, "nombre")
        last_name = _cell(row_values, "apellido")
        if not first_name:
            errors.append("El nombre es obligatorio.")
        elif len(first_name) > 150:
            errors.append("El nombre no puede superar 150 caracteres.")
        if not last_name:
            errors.append("El apellido es obligatorio.")
        elif len(last_name) > 150:
            errors.append("El apellido no puede superar 150 caracteres.")

        # ── Role ─────────────────────────────────────────────────────────────
        role_raw = _cell(row_values, "rol")
        role = _ROLE_MAP.get(role_raw.lower(), None) if role_raw else None
        if not role_raw:
            errors.append("El rol es obligatorio.")
        elif role is None:
            errors.append(
                f"Rol '{role_raw}' no reconocido. Valores aceptados: Administrador, Capacitador, Usuario."
            )

        # ── Optional fields ──────────────────────────────────────────────────
        area_name = _cell(row_values, "area")
        area_id: int | None = None
        if area_name:
            area_id = area_by_name.get(area_name.lower())
            if area_id is None:
                errors.append(
                    f"El área '{area_name}' no existe en el catálogo. "
                    "Créala primero en Configuración → Áreas."
                )
        cargo = _cell(row_values, "cargo")
        grupo_name = _cell(row_values, "grupo")
        grupo_id: int | None = None
        if grupo_name:
            grupo_id = group_by_name.get(grupo_name.lower())
            if grupo_id is None:
                errors.append(f"El grupo '{grupo_name}' no existe.")

        # ── Result ───────────────────────────────────────────────────────────
        if errors:
            error_rows.append(ErrorRow(row=row_number, email=email_raw, errors=errors))
        else:
            assert role is not None  # validated above
            valid_rows.append(
                ValidRow(
                    row=row_number,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    role=role,
                    area_id=area_id,
                    cargo=cargo,
                    grupo_id=grupo_id,
                )
            )

    return BulkImportPreviewResult(
        valid_count=len(valid_rows),
        error_count=len(error_rows),
        valid_rows=valid_rows,
        error_rows=error_rows,
    )


def bulk_import_commit(
    rows: list[ValidRow],
    *,
    admin_user: User,
    ip: str,
) -> BulkImportCommitResult:
    """
    Create users from the previously-previewed valid rows.

    Re-validates email uniqueness to handle race conditions between
    preview and commit. Records a single AuditLog entry.
    """
    created = 0
    commit_errors: list[ErrorRow] = []

    for row_data in rows:
        row_num = row_data["row"]
        email = row_data["email"]
        # Lightweight race-condition guard
        if User.objects.filter(email=email).exists():
            commit_errors.append(
                ErrorRow(
                    row=row_num,
                    email=email,
                    errors=["Ya existe un usuario con este correo (registrado antes de confirmar)."],
                )
            )
            continue
        try:
            create_user(
                email=email,
                first_name=row_data["first_name"],
                last_name=row_data["last_name"],
                role=row_data["role"],
                area_id=row_data.get("area_id"),
                cargo=row_data.get("cargo", ""),
                grupo_id=row_data.get("grupo_id"),
                admin_user=admin_user,
                ip=ip,
            )
            created += 1
        except Exception:  # noqa: BLE001
            commit_errors.append(
                ErrorRow(row=row_num, email=email, errors=["Error interno al crear el usuario."])
            )

    AuditLog.objects.create(
        user=admin_user,
        accion="BULK_USER_IMPORT",
        ip=ip,
        detalles_json={
            "created": created,
            "failed": len(commit_errors),
            "total_rows": len(rows),
        },
    )

    return BulkImportCommitResult(
        created=created,
        failed=len(commit_errors),
        errors=commit_errors,
    )
